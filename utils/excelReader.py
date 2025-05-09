import openpyxl


def read_test_data(file_path, sheet_name):
    """
    Reads data from an Excel sheet and returns a list of tuples (excluding header).
    Assumes the first row is the header.
    """
    wb = openpyxl.load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows or len(rows) < 2:
        raise ValueError("Sheet must have at least one header and one data row.")

    return [row for row in rows[1:] if any(row)]
